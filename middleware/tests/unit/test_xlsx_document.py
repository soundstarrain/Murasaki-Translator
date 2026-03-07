from pathlib import Path

import pytest
from openpyxl import load_workbook, Workbook

from murasaki_translator.core.chunker import TextBlock
from murasaki_translator.documents.xlsx import XlsxDocument


def _build_workbook(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "原文1"
    worksheet["A2"] = "原文2"
    worksheet["A3"] = 123
    workbook.save(path)
    workbook.close()
    return path


@pytest.mark.unit
def test_xlsx_document_line_mode_load_and_save(tmp_path: Path):
    src = _build_workbook(tmp_path / "input.xlsx")
    out = tmp_path / "output.xlsx"

    doc = XlsxDocument(str(src)).set_runtime_context(
        engine_mode="v2",
        chunk_type="line",
        document_role="input",
    )
    items = doc.load()

    assert [item["text"] for item in items] == ["原文1", "原文2"]
    assert [item["meta"] for item in items] == [0, 1]

    blocks = [
        TextBlock(id=1, prompt_text="译文1", metadata=[0]),
        TextBlock(id=2, prompt_text="译文2", metadata=[1]),
    ]
    doc.save(str(out), blocks)

    workbook = load_workbook(out)
    worksheet = workbook["Sheet1"]
    assert worksheet["A1"].value == "原文1"
    assert worksheet["B1"].value == "译文1"
    assert worksheet["B2"].value == "译文2"
    workbook.close()


@pytest.mark.unit
def test_xlsx_document_output_role_reads_target_cells(tmp_path: Path):
    src = _build_workbook(tmp_path / "resume.xlsx")
    workbook = load_workbook(src)
    worksheet = workbook["Sheet1"]
    worksheet["B1"] = "已有译文"
    workbook.save(src)
    workbook.close()

    doc = XlsxDocument(str(src)).set_runtime_context(
        engine_mode="v2",
        chunk_type="line",
        document_role="output",
    )
    items = doc.load()

    assert items[0]["text"] == "已有译文"
    assert items[0]["meta"] == 0
    assert items[1]["text"] == ""


@pytest.mark.unit
def test_xlsx_document_block_mode_load_and_plain_fallback_save(tmp_path: Path):
    src = _build_workbook(tmp_path / "block.xlsx")
    out = tmp_path / "block_out.xlsx"

    doc = XlsxDocument(str(src)).set_runtime_context(
        engine_mode="v1",
        chunk_type="block",
        document_role="input",
    )
    items = doc.load()

    assert len(items) == 2
    assert items[0]["text"].startswith("@id=1@")
    assert items[1]["meta"]["record_id"] == 2

    translated_blocks = [
        TextBlock(
            id=1,
            prompt_text="译文1\n译文2",
            metadata=[items[0]["meta"], items[1]["meta"]],
        )
    ]
    doc.save(str(out), translated_blocks)

    workbook = load_workbook(out)
    worksheet = workbook["Sheet1"]
    assert worksheet["B1"].value == "译文1"
    assert worksheet["B2"].value == "译文2"
    workbook.close()


@pytest.mark.unit
def test_xlsx_document_block_mode_missing_translation_falls_back_to_source(tmp_path: Path):
    src = _build_workbook(tmp_path / "fallback.xlsx")
    out = tmp_path / "fallback_out.xlsx"

    doc = XlsxDocument(str(src)).set_runtime_context(
        engine_mode="v1",
        chunk_type="block",
        document_role="input",
    )
    doc.load()
    translated_blocks = [
        TextBlock(id=1, prompt_text="", metadata=[]),
    ]
    doc.save(str(out), translated_blocks)

    workbook = load_workbook(out)
    worksheet = workbook["Sheet1"]
    assert worksheet["B1"].value == "原文1"
    assert worksheet["B2"].value == "原文2"
    workbook.close()
