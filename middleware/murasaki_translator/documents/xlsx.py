"""XLSX Document Handler - Supports spreadsheet translation with shared V1/V2 abstractions."""

from __future__ import annotations

import logging
import re
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised via runtime guard
    load_workbook = None

from .base import BaseDocument
from murasaki_translator.core.anchor_guard import (
    normalize_anchor_stream,
    repair_and_validate_anchor_output,
)
from murasaki_translator.core.chunker import TextBlock
from murasaki_translator.utils.alignment_handler import AlignmentHandler

logger = logging.getLogger("murasaki.document.xlsx")


class XlsxDocument(BaseDocument):
    ALIGNMENT_META_KIND = "alignment_structural"

    def __init__(self, path: str):
        super().__init__(path)
        self.records: List[Dict[str, Any]] = []
        self._records_loaded = False

    def load(self) -> List[Dict[str, Any]]:
        self._ensure_dependency()
        self._ensure_records_loaded()

        role = str(self.get_runtime_option("document_role", "input") or "input").strip().lower()
        engine_mode = str(self.get_runtime_option("engine_mode", "") or "").strip().lower()
        chunk_type = str(self.get_runtime_option("chunk_type", "block") or "block").strip().lower()
        use_tagged = role == "input" and (engine_mode == "v1" or chunk_type == "block")

        items: List[Dict[str, Any]] = []
        for index, record in enumerate(self.records):
            if role == "output":
                text = record.get("target_text") or ""
                items.append({"text": text, "meta": index})
                continue

            source_text = record.get("source_text") or ""
            if use_tagged:
                items.append(
                    {
                        "text": f"@id={record['id']}@ {source_text} @end={record['id']}@\n\n",
                        "meta": {
                            "kind": self.ALIGNMENT_META_KIND,
                            "record_id": record["id"],
                            "record_index": index,
                        },
                    }
                )
            else:
                items.append({"text": source_text, "meta": index})
        return items

    def save(self, output_path: str, blocks: List[TextBlock]):
        self._ensure_dependency()
        self._ensure_records_loaded()

        engine_mode = str(self.get_runtime_option("engine_mode", "") or "").strip().lower()
        chunk_type = str(self.get_runtime_option("chunk_type", "block") or "block").strip().lower()
        use_line_save = engine_mode == "v2" and chunk_type == "line"
        if use_line_save:
            value_map = self._build_line_value_map(blocks)
        else:
            value_map = self._build_block_value_map(blocks)
        self._write_output_workbook(output_path, value_map)

    def _ensure_dependency(self) -> None:
        if load_workbook is None:
            raise RuntimeError(
                "XLSX support requires openpyxl. Please install dependency 'openpyxl'."
            )

    def _ensure_records_loaded(self) -> None:
        if self._records_loaded:
            return
        self.records = self._scan_records()
        self.raw_lines = [record["source_text"] for record in self.records]
        self.metadata["record_count"] = len(self.records)
        self._records_loaded = True

    @staticmethod
    def _normalize_cell_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).replace("\r\n", "\n").replace("\r", "\n")
        parts = [segment.strip() for segment in text.split("\n") if segment.strip()]
        return " ".join(parts).strip()

    @staticmethod
    def _is_translatable_source(text: str) -> bool:
        if not text:
            return False
        if text.startswith("="):
            return False
        if re.fullmatch(r"[0-9０-９\-_.:/\\]+", text):
            return False
        return True

    def _scan_records(self) -> List[Dict[str, Any]]:
        workbook = load_workbook(self.path)
        records: List[Dict[str, Any]] = []
        record_id = 1
        try:
            for worksheet in workbook.worksheets:
                max_col = max(1, int(worksheet.max_column or 1))
                max_row = int(worksheet.max_row or 0)
                for row_idx in range(1, max_row + 1):
                    source_col: Optional[int] = None
                    source_text = ""
                    for col_idx in range(1, max_col + 1):
                        candidate = self._normalize_cell_text(
                            worksheet.cell(row=row_idx, column=col_idx).value
                        )
                        if not self._is_translatable_source(candidate):
                            continue
                        source_col = col_idx
                        source_text = candidate
                        break
                    if source_col is None:
                        continue

                    target_col = source_col + 1
                    target_text = self._normalize_cell_text(
                        worksheet.cell(row=row_idx, column=target_col).value
                    )
                    records.append(
                        {
                            "id": record_id,
                            "sheet_name": worksheet.title,
                            "row": row_idx,
                            "source_col": source_col,
                            "target_col": target_col,
                            "source_text": source_text,
                            "target_text": target_text,
                        }
                    )
                    record_id += 1
        finally:
            workbook.close()
        return records

    def _build_line_value_map(self, blocks: List[TextBlock]) -> Dict[int, str]:
        translated_by_index: Dict[int, str] = {}
        for block_index, block in enumerate(blocks):
            mapped_index = self._resolve_line_record_index(block, block_index)
            if mapped_index is None or mapped_index < 0 or mapped_index >= len(self.records):
                continue
            text = str(getattr(block, "prompt_text", "") or "").strip("\r\n")
            if not text:
                text = self.records[mapped_index]["source_text"]
            translated_by_index[mapped_index] = text

        return {
            record["id"]: translated_by_index.get(index, record["source_text"])
            for index, record in enumerate(self.records)
        }

    def _build_block_value_map(self, blocks: List[TextBlock]) -> Dict[int, str]:
        expected_ids = [int(record["id"]) for record in self.records]
        full_stream = "\n".join(str(getattr(block, "prompt_text", "") or "") for block in blocks)
        full_stream = normalize_anchor_stream(full_stream)
        expected_source = "\n".join(
            f"@id={record_id}@\n@end={record_id}@" for record_id in expected_ids
        )
        repaired_stream, repair_ok, repair_meta = repair_and_validate_anchor_output(
            expected_source,
            full_stream,
            mode="alignment",
        )
        if repair_meta.get("repaired"):
            logger.info(
                "[XlsxDocument] Alignment stream repaired: %s",
                ",".join(repair_meta.get("repair_steps", [])),
            )
        if not repair_ok:
            logger.warning(
                "[XlsxDocument] Repair validation incomplete: missing=%s strict_missing=%s foreign=%s",
                repair_meta.get("missing_count", 0),
                repair_meta.get("strict_pair_missing_count", 0),
                repair_meta.get("foreign_count", 0),
            )

        logical_trans_map = AlignmentHandler._extract_logical_map(
            repaired_stream,
            expected_ids=expected_ids,
        )
        missing_ids = [record_id for record_id in expected_ids if record_id not in logical_trans_map]
        if missing_ids:
            self._fill_missing_from_block_plain_text(
                logical_trans_map,
                missing_ids,
                blocks,
            )

        source_lookup = {
            int(record["id"]): str(record.get("source_text") or "")
            for record in self.records
        }
        for record_id in expected_ids:
            if record_id not in logical_trans_map:
                logical_trans_map[record_id] = source_lookup.get(record_id, "")

        return logical_trans_map

    def _fill_missing_from_block_plain_text(
        self,
        logical_trans_map: Dict[int, str],
        missing_ids: List[int],
        translated_blocks: List[TextBlock],
    ) -> int:
        missing_set = set(missing_ids)
        assigned = 0

        for block in translated_blocks:
            if not missing_set:
                break

            block_ids = [
                record_id
                for record_id in self._extract_record_ids_from_metadata(block.metadata)
                if record_id in missing_set
            ]
            if not block_ids:
                continue

            block_text = normalize_anchor_stream(str(getattr(block, "prompt_text", "") or ""))
            if not block_text.strip():
                continue

            block_map = AlignmentHandler._extract_logical_map(
                block_text,
                expected_ids=block_ids,
            )
            for record_id in block_ids:
                content = block_map.get(record_id)
                if content is None or record_id not in missing_set:
                    continue
                logical_trans_map[record_id] = content
                missing_set.remove(record_id)
                assigned += 1

            unresolved_ids = [record_id for record_id in block_ids if record_id in missing_set]
            if not unresolved_ids:
                continue

            all_block_ids = self._extract_record_ids_from_metadata(block.metadata)
            plain_block = AlignmentHandler.TAG_MARKER_PATTERN.sub("", block_text).strip()
            plain_lines = [line.strip() for line in plain_block.splitlines() if line.strip()]
            if not plain_lines:
                continue

            if len(unresolved_ids) == 1:
                unresolved_id = unresolved_ids[0]
                candidate = ""
                if len(all_block_ids) == 1:
                    candidate = " ".join(plain_lines).strip()
                elif len(plain_lines) == len(all_block_ids):
                    unresolved_pos = all_block_ids.index(unresolved_id)
                    if 0 <= unresolved_pos < len(plain_lines):
                        candidate = plain_lines[unresolved_pos]
                elif len(block_ids) == 1 and len(plain_lines) == 1:
                    candidate = plain_lines[0]

                if candidate:
                    logical_trans_map[unresolved_id] = candidate
                    missing_set.remove(unresolved_id)
                    assigned += 1
                continue

            if len(plain_lines) == len(unresolved_ids):
                for record_id, line in zip(unresolved_ids, plain_lines):
                    logical_trans_map[record_id] = line
                    missing_set.remove(record_id)
                    assigned += 1

        if assigned > 0:
            logger.warning(
                "[XlsxDocument] Filled %s missing logical IDs via block-level fallback.",
                assigned,
            )
        return assigned

    @staticmethod
    def _resolve_line_record_index(block: TextBlock, default_index: int) -> Optional[int]:
        for meta in getattr(block, "metadata", None) or []:
            if isinstance(meta, int):
                return meta
            if isinstance(meta, dict):
                record_index = meta.get("record_index")
                if isinstance(record_index, int):
                    return record_index
        return default_index

    @classmethod
    def _extract_record_ids_from_metadata(cls, metadata: Optional[List[Any]]) -> List[int]:
        ordered: List[int] = []
        seen = set()
        for item in metadata or []:
            record_id: Optional[int] = None
            if isinstance(item, dict) and item.get("kind") == cls.ALIGNMENT_META_KIND:
                candidate = item.get("record_id")
                if isinstance(candidate, int):
                    record_id = candidate
                elif candidate is not None:
                    try:
                        record_id = int(candidate)
                    except (TypeError, ValueError):
                        record_id = None
            if record_id is None or record_id in seen:
                continue
            seen.add(record_id)
            ordered.append(record_id)
        return ordered

    def _write_output_workbook(self, output_path: str, value_map: Dict[int, str]) -> None:
        workbook = load_workbook(self.path)
        try:
            for record in self.records:
                worksheet = workbook[record["sheet_name"]]
                worksheet.cell(
                    row=int(record["row"]),
                    column=int(record["target_col"]),
                    value=value_map.get(int(record["id"]), record["source_text"]),
                )
            workbook.save(output_path)
        finally:
            workbook.close()
